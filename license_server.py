# license_server.py (final, disk-enabled version with secure installer download)

from flask import Flask, request, send_file, jsonify, render_template_string, redirect
import os
import json
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# ✅ Correct disk-mounted paths (Render Starter Plan with /mnt/data)
BASE_DIR = "/mnt/data"
TEMPLATE_FILE = os.path.join(os.path.dirname(__file__), "template.xlsm")
INSTALLER_FILE = os.path.join(os.path.dirname(__file__), "installer_lifetime.exe")
ALLOWED_FILE = os.path.join(BASE_DIR, "allowed_ids.json")
PENDING_FILE = os.path.join(BASE_DIR, "pending_ids.json")
OUTPUT_DIR = BASE_DIR

# Utility: Ensure files exist
for path in [ALLOWED_FILE, PENDING_FILE]:
    if not os.path.exists(path):
        with open(path, "w") as f:
            json.dump([], f)

def load_ids(path):
    with open(path, "r") as f:
        return json.load(f)

def save_ids(path, ids):
    with open(path, "w") as f:
        json.dump(ids, f, indent=2)

def generate_password(machine_id):
    seed = 12345
    for c in machine_id:
        seed += ord(c)
    return f"PWD{seed}"

@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    machine_id = data.get("machine_id", "").strip().upper()
    if not machine_id:
        return jsonify({"error": "Missing machine ID"}), 400

    allowed = load_ids(ALLOWED_FILE)
    if machine_id not in allowed:
        return jsonify({"error": "Machine ID not approved."}), 403

    if not os.path.exists(TEMPLATE_FILE):
        return jsonify({"error": "Template file not found."}), 500

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_file = os.path.join(OUTPUT_DIR, f"QTY_Network_2025_{timestamp}.xlsm")

    wb = load_workbook(TEMPLATE_FILE, keep_vba=True)
    ws = wb["LicenseData"]
    ws["A1"] = machine_id
    ws["A2"] = generate_password(machine_id)
    wb.save(output_file)

    return send_file(output_file, as_attachment=True)

@app.route("/request", methods=["POST"])
def request_access():
    data = request.get_json()
    machine_id = data.get("machine_id", "").strip().upper()
    if not machine_id:
        return jsonify({"error": "Missing machine ID"}), 400

    pending = load_ids(PENDING_FILE)
    allowed = load_ids(ALLOWED_FILE)

    if machine_id in pending or machine_id in allowed:
        return jsonify({"status": "Already received or approved."})

    pending.append(machine_id)
    save_ids(PENDING_FILE, pending)
    return jsonify({"status": "Request received. Waiting for approval."})

@app.route("/admin")
def admin():
    pending = load_ids(PENDING_FILE)
    html = """
    <h2>Pending Machine ID Requests</h2>
    <ul>
    {% for mid in pending %}
      <li>{{ mid }} ✅ <a href='/approve/{{ mid }}'>Approve</a> ❌ <a href='/reject/{{ mid }}'>Reject</a></li>
    {% endfor %}
    </ul>
    """
    return render_template_string(html, pending=pending)

@app.route("/approve/<machine_id>")
def approve(machine_id):
    machine_id = machine_id.strip().upper()
    allowed = load_ids(ALLOWED_FILE)
    pending = load_ids(PENDING_FILE)

    if machine_id not in allowed:
        allowed.append(machine_id)
    if machine_id in pending:
        pending.remove(machine_id)

    save_ids(ALLOWED_FILE, allowed)
    save_ids(PENDING_FILE, pending)
    return redirect("/admin")

@app.route("/reject/<machine_id>")
def reject(machine_id):
    machine_id = machine_id.strip().upper()
    pending = load_ids(PENDING_FILE)

    if machine_id in pending:
        pending.remove(machine_id)
        save_ids(PENDING_FILE, pending)

    return redirect("/admin")

@app.route("/download")
def download():
    machine_id = request.args.get("mid", "").strip().upper()
    allowed = load_ids(ALLOWED_FILE)

    if machine_id in allowed:
        if os.path.exists(INSTALLER_FILE):
            return send_file(INSTALLER_FILE, as_attachment=True)
        else:
            return "❌ Installer not found on server.", 404
    else:
        return "❌ Access Denied. This machine is not approved.", 403

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=10000)
